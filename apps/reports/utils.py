"""
Утилиты для генерации отчётов.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.utils import timezone
import datetime


def generate_equipment_excel(equipment_queryset):
    """
    Генерирует отчёт по оборудованию в формате Excel.
    
    Args:
        equipment_queryset: QuerySet с объектами оборудования
        
    Returns:
        bytes: Буфер с данными Excel файла
    """
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборудование"
    
    # Заголовки
    headers = [
        'Инвентарный номер',
        'Наименование',
        'Категория',
        'Статус',
        'Серийный номер',
        'Производитель',
        'Модель',
        'Дата покупки',
        'Цена покупки',
        'Подразделение',
        'Ответственное лицо',
        'Местоположение'
    ]
    
    # Стиль заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Запись заголовков
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Запись данных
    for row_idx, equipment in enumerate(equipment_queryset, 2):
        data = [
            equipment.inventory_number,
            equipment.name,
            equipment.category.name if equipment.category else '',
            equipment.get_status_display(),
            equipment.serial_number or '',
            equipment.manufacturer or '',
            equipment.model or '',
            equipment.purchase_date.strftime('%d.%m.%Y') if equipment.purchase_date else '',
            equipment.purchase_price or '',
            equipment.department.name if equipment.department else '',
            equipment.responsible_person.get_full_name() if equipment.responsible_person else '',
            equipment.location or ''
        ]
        
        for col_idx, value in enumerate(data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Автоматическая ширина колонок
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в буфер
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output